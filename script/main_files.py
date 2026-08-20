# -- coding: utf-8 --
"""
Автор: Олег Шабалов
Контакты: 89179021656, @olegshabalov

Выгрузка данных из ВБ в базу
"""

import logging
import shutil
import sys
import os

import openpyxl

from db import Database
import lib
from telegram_bot.notifier import notify_success, notify_error

from settings import LOCAL_MODE, DIR_DATA, WB_DATE_FORMAT, WB_DATETIME_FORMAT, PSQL_DATABASE, PSQL_USER, PSQL_PASW, PSQL_HOST, \
    PSQL_PORT

FILES_PATH = 'FILES' if LOCAL_MODE else '/home/user/FILES'
CSV_DATE_FORMAT = "%d.%m.%Y"


def read_csv_ozon_product_advert(file_path):
    delimiter = ';'
    encoding = 'utf8'
    res = []
    header_dict = {}
    dt = None
    try:
        with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
            for i, line in enumerate(f):
                line = line.strip()
                if i == 0:
                    dt_list = line.replace('Период', '').replace('\ufeff', '').replace(' ', '').split('-')
                elif i < 2:
                    continue
                elif i == 2:
                    header_dict = {index: key.replace('₽', '').replace('%', '').replace(',', '').replace('  ', ' ').strip().replace(' ', '_') for index, key in enumerate(line.split(delimiter))}
                else:
                    res_elem = {header_dict.get(index): elem for index, elem in enumerate(line.split(delimiter))}
                    res.append(res_elem)
    except Exception as e:
        logging.warning(str(e))
    return res, dt_list


def read_xlsx_ozon_product_advert(filepath):
    dt_list = []
    res = []
    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
    except Exception as e:
        logging.warning(str(e))
        logging.warning(f'Ошибка открытия файла {filepath}')
    else:
        row_num = 0
        headers_dict = dict()
        for row in sheet.iter_rows():
            row_num += 1
            if row_num == 1:
                if row[0]:
                    dt_list = str(row[0].value).replace('Период:', '').replace('\ufeff', '').replace(' ', '').split('-')
            elif row_num == 2:
                for i, cell in enumerate(row):
                    if cell.value:
                        headers_dict[i] = cell.value.replace('₽', '').replace('%', '').replace(',', '').replace('  ', ' ').strip().replace(' ', '_')
                continue

            elem_dict = dict()
            for i, cell in enumerate(row):
                if i in headers_dict:
                    elem_dict[headers_dict[i]] = str(cell.value).strip() if cell.value else ''

            if elem_dict:
                res.append(elem_dict)
    return res, dt_list


def read_xlsx_ozon_product_advert2(filepath):
    float_cols_indexes = [5, 6, 7, 9, 13, 14, 15]
    int_cols_indexes = [4, 8, 10, 11, 12]
    dt_list = []
    res = []
    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
    except Exception as e:
        logging.warning(str(e))
        logging.warning(f'Ошибка открытия файла {filepath}')
    else:
        row_num = 0
        headers_dict = dict()
        for row in sheet.iter_rows():
            row_num += 1
            if row_num == 1:
                if row[0]:
                    dt_list = str(row[0].value).replace('Период:', '').replace('\ufeff', '').replace(' ', '').split('-')
            elif row_num == 2:
                for i, cell in enumerate(row):
                    if cell.value:
                        headers_dict[i] = cell.value.replace('₽', '').replace('%', '').replace(',', '').replace('  ', ' ').strip().replace(' ', '_')
                continue

            elem_dict = dict()
            for i, cell in enumerate(row):
                if i in headers_dict:
                    value = str(cell.value).strip() if cell.value else ''
                    if i in float_cols_indexes:
                        try:
                            value = float(value.replace(',', '.'))
                        except:
                            value = 0.0
                    elif i in int_cols_indexes:
                        try:
                            value = int(float(value.replace(',', '.')))
                        except:
                            value = 0
                    elem_dict[headers_dict[i]] = value

            if elem_dict:
                res.append(elem_dict)
    return res, dt_list


def read_xlsx_wb_media(filepath):
    dt_list = []
    res = []
    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        filename_list = filepath.rstrip('.xlsx').split('-')
        dt_list = filename_list[-2:]
    except Exception as e:
        logging.warning(str(e))
        logging.warning(f'Ошибка открытия файла {filepath}')
    else:
        row_num = 0
        headers_dict = dict()
        for row in sheet.iter_rows():
            row_num += 1
            if row_num == 1:
                for i, cell in enumerate(row):
                    if cell.value:
                        headers_dict[i] = cell.value.replace('₽', '').replace('%', '').replace(',', '').replace('  ', ' ').strip().replace(' ', '_')
                continue

            elem_dict = dict()
            for i, cell in enumerate(row):
                if i in headers_dict:
                    elem_dict[headers_dict[i]] = str(cell.value).strip() if cell.value else ''

            if elem_dict:
                # elem_dict['SKU'] = 387558556 #FIXME!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
                res.append(elem_dict)
    return res, dt_list


####################################################### ENTITIES #######################################################
ENTITIES = [
    # {
    #     'table': 'ozon_advert_csv',
    #     'index_fields': ['client_id', 'date', 'SKU', 'ID_кампании'],
    #     'func': read_csv_ozon_product_advert,
    #     'skip_fields': None,
    #     'file_types': ['csv']
    # },
    {
        'table': 'ozon_advert_csv',
        'dir': 'ozon_advert_csv',
        'index_fields': ['client_id', 'date', 'SKU', 'ID_кампании'],
        'func': read_xlsx_ozon_product_advert,
        'skip_fields': None,
        'file_types': ['xlsx'],
        'clear': False
    },
    {
        'table': 'ozon_advert_csv2',
        'dir': 'ozon_advert_csv',
        'index_fields': ['client_id', 'date', 'SKU', 'ID_кампании'],
        'func': read_xlsx_ozon_product_advert2,
        'skip_fields': None,
        'file_types': ['xlsx'],
        'clear': True
    },
    {
        'table': 'wb_media_adverts_xlsx',
        'dir': 'wb_media',
        'index_fields': ['client_id', 'date', 'SKU', 'CampaignID'],
        'func': read_xlsx_wb_media,
        'skip_fields': None,
        'file_types': ['xlsx'],
        'clear': True
    },

]
########################################################################################################################


if __name__ == "__main__":
    script_file_name = os.path.basename(__file__)
    logging.basicConfig(handlers=[logging.FileHandler(filename=os.path.join(DIR_DATA, f"log_{script_file_name.split('.')[0]}.txt"), encoding='utf-8', mode='a+')],
                        format='[%(asctime)s] [%(levelname)s] => %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S', level=logging.INFO)
    logging.info("-" * 50)

    def _excepthook(exc_type, exc_value, exc_tb):
        logging.critical("Unhandled exception", exc_info=(exc_type, exc_value, exc_tb))
        notify_error(script_file_name, exc_value)
        sys.__excepthook__(exc_type, exc_value, exc_tb)

    sys.excepthook = _excepthook

    logging.info(f'Script {script_file_name} START')

    db = Database(PSQL_DATABASE, PSQL_USER, PSQL_PASW, PSQL_HOST, PSQL_PORT)
    if not db.conn:
        logging.warning("DB connection error - FINISH")
        notify_error(script_file_name, "DB connection error")
        sys.exit()

    for entity in ENTITIES:
        func = entity.get('func')
        table = entity.get('table')
        dir = entity.get('dir')
        clear = entity.get('clear')
        index_fields = entity.get('index_fields', [])
        if not index_fields:
            index_fields = ['id']
        norm_index_fields = [lib.camel_to_snake(index_field) for index_field in index_fields]
        agregate_index_field = '_'.join(norm_index_fields)
        agregate_index_fields = [agregate_index_field]
        agregate_index_field_set = set()

        if not func or not table or not dir:
            continue

        clients_folders = [f.path for f in os.scandir(os.path.join(FILES_PATH, dir)) if f.is_dir()]
        for client_folder in clients_folders:
            logging.info(f"Process entity {table}, client_folder {client_folder}")
            client_id = client_folder.split('_')[-1]
            try:
                int(client_id)
            except:
                logging.warning(f"Couldnt get client_id from {client_folder}")
                continue
            file_paths = [f for f in os.listdir(client_folder) if os.path.isfile(os.path.join(client_folder, f)) and f.split('.')[-1].lower() in entity.get('file_types', [])]
            for file_name in file_paths:
                file_path = os.path.join(client_folder, file_name)
                elems, dt_list = func(file_path)
                # print(str(elems).encode('cp1251', 'ignore').decode('cp1251')) #FIXME
                # print(dt_list)
                # continue #FIXME
                logging.info(f"Recieved {len(elems)} items from {file_path}")

                if not elems:
                    ########################################## move to arhive ##########################################
                    try:
                        arhive_path = os.path.join(client_folder, 'arhive')
                        if not os.path.exists(arhive_path):
                            os.makedirs(arhive_path)
                        shutil.copyfile(file_path, os.path.join(arhive_path, file_name))
                        os.remove(file_path)
                    except Exception as e:
                        logging.warning(str(e))
                    ####################################################################################################
                    continue

                if len(dt_list) != 2 or dt_list[0] != dt_list[1]:
                    logging.info(f"Дата {dt_list} - ОШИБКА!")
                    # continue $FIXME

                if not db.test():
                    db.reconnect()

                flag = True
                add_flag = True
                template_fields_dict = {}
                i = 0
                err = 0
                for elem in elems:
                    elem['client_id'] = client_id
                    elem['date'] = dt_list[0]

                    if index_fields == ['id']:
                        elem['id'] = i + 1
                    stop_flag = False
                    agregate_index_value = ''
                    for index_field in index_fields:
                        if not elem.get(index_field):
                            stop_flag = True
                    if stop_flag:
                        logging.warning(f"Row {elem} no index fields {index_fields}")
                        continue

                    flat_dict = {}
                    lib.normalize_dict(elem, flat_dict, '', skip_fields=entity.get('skip_fields'), date_format=CSV_DATE_FORMAT, datetime_format=WB_DATETIME_FORMAT)
                    if agregate_index_field not in flat_dict:
                        flat_dict[agregate_index_field] = '_'.join([str(flat_dict.get(index_key)) for index_key in norm_index_fields])
                    agregate_index_field_set.add(flat_dict[agregate_index_field])

                    if flag:
                        init_template_fields_dict = lib.make_pg_dict_template(flat_dict)
                        # print('our', template_fields_dict)
                        agregate_index_fields = db.create_table_from_dict_template(table, init_template_fields_dict, index_fields=agregate_index_fields)
                        if not agregate_index_fields:
                            logging.warning(f"Error create_table_from_dict_template")
                            break
                        template_fields_dict = db.get_fields_dict(table)
                        flag = False

                        fields_diff = init_template_fields_dict.keys() - template_fields_dict.keys()
                        if fields_diff:
                            for field in fields_diff:
                                field_type = lib.recoginze_pg_value_type(flat_dict[field])
                                if db.add_table_column(table, field, field_type):
                                    template_fields_dict[field] = field_type
                                else:
                                    add_flag = False
                                    logging.warning(f"Error add_table_column")
                                    break
                    else:
                        fields_dict = lib.make_pg_dict_template(flat_dict)
                        fields_diff = fields_dict.keys() - template_fields_dict.keys()
                        if fields_diff:
                            for field in fields_diff:
                                field_type = lib.recoginze_pg_value_type(flat_dict[field])
                                if db.add_table_column(table, field, field_type):
                                    template_fields_dict[field] = field_type
                                else:
                                    add_flag = False
                                    logging.warning(f"Error add_table_column")
                                    break
                    # print(template_fields_dict)
                    if not add_flag:
                        add_flag = True
                        logging.warning(f"Error add_flag")
                        err += 1
                        continue

                    if db.insert_update_item(table, flat_dict, agregate_index_fields):
                        i += 1
                logging.info(f"Totaly insert/update {i} (errors {err}) in table {table}")
                logging.info(f"Totaly {len(agregate_index_field_set)} uniq index values in table {table}")

                ########################################## move to arhive ##########################################
                if clear:
                    try:
                        arhive_path = os.path.join(client_folder, 'arhive')
                        if not os.path.exists(arhive_path):
                            os.makedirs(arhive_path)
                        shutil.copyfile(file_path, os.path.join(arhive_path, file_name))
                        os.remove(file_path)
                    except Exception as e:
                        logging.warning(str(e))
        logging.info('-'*100)

    db.close()
    logging.info('FINISH')
    notify_success(script_file_name)
    sys.exit()
