# -- coding: utf-8 --
"""
Автор: Олег Шабалов
Контакты: 89179021656, @olegshabalov

Библиотека общих функций
"""

import datetime
import logging
import json
import re
import sys
from zipfile import ZipFile

import openpyxl
import requests
from transliterate import translit

from settings import MAX_STR_SIZE

HEADERS = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'}


def read_xlsx_to_dict(file_path, header_row=1, sub_header_row=None, values_row=None, sheets=None) -> dict[
    str, list[dict]]:
    """
    Читает эксель в словарь с ключами = названия листов, а в них словари с ключами = заголовки столбцов
    """
    res = dict()
    wb = openpyxl.load_workbook(file_path)
    sheetnames = wb.sheetnames
    for sheetname in sheetnames:
        if sheets and sheetname not in sheets:
            continue
        sheet = wb[sheetname]
        headers = dict()
        sheet_lat_name = sheetname
        res[sheet_lat_name] = []
        for i in range(sheet.max_column):
            headers[i] = sheet.cell(row=header_row, column=i + 1).value
            if not headers[i]:
                headers[i] = ''

        if sub_header_row:
            for i in range(sheet.max_column):
                if headers[i]:
                    headers[i] += '_'
                if len(headers[i]) > 50:
                    headers[i] = ''
                headers[i] += sheet.cell(row=sub_header_row, column=i + 1).value
            if not values_row:
                values_row = sub_header_row + 1

        if not values_row:
            values_row = header_row + 1

        for j in range(values_row - 1, sheet.max_row):
            elem = dict()
            add_flag = False
            for i in range(sheet.max_column):
                value = sheet.cell(row=j + 1, column=i + 1).value
                if value:
                    add_flag = True
                elem[headers[i]] = value
            if add_flag:
                res[sheet_lat_name].append(elem)

    return res


def camel_to_snake(name):
    name = re.sub('(.)([A-Z][a-z]+)', r'\1_\2', name)
    return re.sub('([a-z0-9])([A-Z])', r'\1_\2', name).lower()


def anti_vowel(text):
    text = list(text)
    for i in text[::-1]:
        if i in 'aeiouAEIOU':
            text.remove(i)
    return str(''.join(text))


def normalize_dict(src, dst_dict: dict, parent_name: str, skip_fields: list = None, date_format: str = None,
                   datetime_format: str = None, snake_case=True, limit_list_fields: list = None) -> None:
    """
    For psql snake_case=True
    For mysql snake_case=False
    """
    length_limit = 64

    def test_convert_date(value, date_format=None, datetime_format=None):
        if datetime_format and len(value) >= len(datetime_format):
            try:
                res = datetime.datetime.strptime(value.replace('Z', '').replace(' ', 'T').split('.')[0].split('+')[0],
                                                 datetime_format)
            except:
                return value
        elif date_format:
            try:
                res = datetime.datetime.strptime(value.replace('Z', '').split('.')[0].split('+')[0], date_format)
            except:
                return value
        else:
            return value
        return res

    if snake_case:
        parent_name = camel_to_snake(parent_name)
    parent_name = parent_name.replace('.', '_')

    if not src:
        return
    if not isinstance(src, dict):
        if src is None:
            src = ''
        if len(parent_name) > length_limit:
            parent_name = anti_vowel(parent_name)[:length_limit]
        if date_format and isinstance(src, str):
            src = test_convert_date(src, date_format, datetime_format)
        dst_dict[parent_name] = src
        return
    for key in src:
        if skip_fields and key in skip_fields:
            continue
        if isinstance(src[key], dict):
            new_key = key.replace(' ', '')
            if parent_name:
                new_key = parent_name + '_' + new_key
            normalize_dict(src[key], dst_dict, new_key, skip_fields=skip_fields, date_format=date_format,
                           datetime_format=datetime_format)
        elif isinstance(src[key], list):
            new_key = key.replace(' ', '')
            if parent_name:
                new_key = parent_name + '_' + new_key
            for i, list_elem in enumerate(src[key]):
                normalize_dict(list_elem, dst_dict, new_key + f"_{i}", skip_fields=skip_fields, date_format=date_format,
                               datetime_format=datetime_format)
                if limit_list_fields and key in limit_list_fields:
                    break
        else:
            new_key = key.replace(' ', '')
            if snake_case:
                new_key = camel_to_snake(new_key)
            new_key = new_key.replace('.', '_')
            if parent_name:
                new_key = parent_name + '_' + new_key
            if len(new_key) > length_limit:
                new_key = anti_vowel(new_key)[:length_limit]

            if src[key] is None:
                value = ''
            else:
                value = src[key]
            if date_format and isinstance(value, str):
                value = test_convert_date(value, date_format, datetime_format)

            dst_dict[new_key] = value


def recoginze_pg_value_type(value) -> str:
    default = f'VARCHAR({MAX_STR_SIZE})'
    if isinstance(value, bool):
        return 'BOOL'
    elif isinstance(value, int):
        return 'BIGINT'
    elif isinstance(value, float):
        return 'DOUBLE PRECISION'
    elif isinstance(value, datetime.datetime):
        return 'TIMESTAMP'
    elif value is None:
        return default
    elif len(value) > MAX_STR_SIZE:
        return 'TEXT'
    return default


def make_pg_dict_template(fields_dict: dict[str, any], preset_fields_types_dict: dict = None) -> dict[str, str]:
    res = dict()
    for key in fields_dict:
        if preset_fields_types_dict and key in preset_fields_types_dict:
            res[key] = preset_fields_types_dict[key]
        else:
            res[key] = recoginze_pg_value_type(fields_dict[key])
    return res


def clear_dict_keys(res):
    ############################################## clear dict ##########################################################
    clear_res = []
    for elem in res:
        new_dict = {}
        for key in elem:
            new_key = str(key).replace(',', '_').replace('.', '_').replace('%', '_').replace(' ', '_').replace('"', '') \
                .replace("'", '').replace("-", '_').replace("₽", 'руб').replace(')', '_').replace('(', '_').replace('/', '_').lower()
            while True:
                if '__' in new_key:
                    new_key = new_key.replace('__', '_')
                else:
                    break
            new_key = translit(new_key.strip('_'), language_code='ru', reversed=True).replace("'", '')
            new_value = elem[key]
            if isinstance(new_value, str):
                new_value = new_value.lstrip("'")
            # print(elem[key], new_value)
            # if new_value.count('"') == 1:
            #     new_value = new_value.replace('"', '')
            new_dict[new_key] = new_value
        clear_res.append(new_dict)

    return clear_res


def read_xlsx_to_dict_v2(file_path, header_row=1, sub_header_row=None, sub_sub_header_row=None, values_row=None, sheets=None, exclude_nons=True) -> dict[
    str, list[dict]]:
    """
    Читает эксель в словарь с ключами = названия листов, а в них словари с ключами = заголовки столбцов
    """
    res = dict()
    wb = openpyxl.load_workbook(file_path)
    sheetnames = wb.sheetnames
    for sheetname in sheetnames:
        if sheets and sheetname not in sheets:
            continue
        sheet = wb[sheetname]
        headers = dict()
        sheet_lat_name = sheetname
        res[sheet_lat_name] = []

        for i in range(sheet.max_column):
            headers[i] = sheet.cell(row=header_row, column=i + 1).value
            if not headers[i] and i != 0:
                headers[i] = headers[i-1]

        if not values_row:
            values_row = header_row + 1

        if sub_header_row:
            for i in range(sheet.max_column):
                sub_header = sheet.cell(row=sub_header_row, column=i + 1).value
                if sub_header:
                    try:
                        headers[i] += '_' + sub_header
                    except Exception as e:
                        logging.warning(f"{str(e)} - {headers[i]}, {sub_header}")
            if not values_row:
                values_row = sub_header_row + 1

        if sub_sub_header_row:
            for i in range(sheet.max_column):
                sub_header = sheet.cell(row=sub_sub_header_row, column=i + 1).value
                if sub_header:
                    try:
                        headers[i] += '_' + sub_header
                    except Exception as e:
                        logging.warning(f"{str(e)} - {headers[i]}, {sub_header}")
            if not values_row:
                values_row = sub_sub_header_row + 1

        for j in range(values_row - 1, sheet.max_row):
            elem = dict()
            for i in range(sheet.max_column):
                value = sheet.cell(row=j + 1, column=i + 1).value
                if exclude_nons and value is None:
                    continue
                elem[headers[i]] = value
            if elem:
                res[sheet_lat_name].append(elem)

    return res


def read_xlsx_to_dict_v3(file_path, sheetname, values_row):
    res = []
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheetname]
    except Exception as e:
        logging.warning(str(e))
        return res
    row = 0
    for elem in sheet.iter_rows(values_only=True):
        row += 1
        if row < values_row:
            continue
        res.append(elem)

    return res


def get_file(url, file_path, headers=HEADERS):
    resp = requests.get(url, headers=headers)
    if not resp.ok:
        logging.warning(f"get_file: {resp.text}")
        return False
    try:
        with open(file_path, 'wb') as f:
            f.write(resp.content)
        return True
    except Exception as e:
        logging.warning(f"get_file:{str(e)}")
        return False


def validate_values_types(flat_dict, template_fields_dict):
    # print(template_fields_dict)
    for field in template_fields_dict:
        # print(field, template_fields_dict[field])
        if template_fields_dict[field].lower() in ['bigint', 'double precision'] and field in flat_dict and not flat_dict.get(field):
            flat_dict[field] = 0
        elif template_fields_dict[field].lower() in ['timestamp', 'timestamp without time zone'] and field in flat_dict and not flat_dict.get(field):
            flat_dict[field] = None
        # if template_fields_dict[field].lower() == 'timestamp':
        #     print(flat_dict[field])
    return flat_dict


def read_json(file, encoding='utf8'):
    try:
        with open(file, encoding=encoding) as f:
            try:
                tmp = json.load(f)
            except Exception as e:
                logging.warning(str(e))
                return {}
    except Exception as e:
        logging.warning(str(e))
        return {}
    return tmp


def unzip_files(arh_path, unzip_dir):
    # unzip_dir = os.path.dirname(arh_path)
    try:
        with ZipFile(arh_path, 'r') as zip_file:
            zip_file.extractall(unzip_dir)
    except Exception as e:
        logging.warning(str(e))
        return False
    return True


def read_csv(file_path, delimiter=',', encoding='utf8'):
    res = []
    header_dict = {}
    try:
        with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
            for i, line in enumerate(f):
                line = line.strip()
                if i == 0:
                    header_dict = {index: key.replace('₽', '').replace('%', '').replace(' ', '_') for index, key in enumerate(line.replace('\ufeff', '').split(delimiter))}
                else:
                    res_elem = {header_dict.get(index): elem for index, elem in enumerate(line.split(delimiter))}
                    res.append(res_elem)
    except Exception as e:
        logging.warning(str(e))
    return res